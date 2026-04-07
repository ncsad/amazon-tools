# -*- coding: utf-8 -*-
import re, os, json, threading
import urllib.request
import numpy as np
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFrame, QLabel, QPushButton,
    QLineEdit, QFileDialog, QHBoxLayout, QVBoxLayout, QGridLayout,
    QSizePolicy, QProgressBar, QGraphicsDropShadowEffect
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QMimeData, QPropertyAnimation, QEasingCurve, QSize
from PyQt6.QtGui import QColor, QFont, QDragEnterEvent, QDropEvent, QPalette, QIcon
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from scipy.optimize import minimize

# ============================================================
# 计算逻辑
# ============================================================

def get_all_rates():
    try:
        url = "https://api.exchangerate-api.com/v4/latest/USD"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            rates = json.loads(r.read()).get("rates", {})
            usd_cny = rates.get("CNY", 7.2)
            gbp_usd = rates.get("GBP", 0.78)
            return {"USD": usd_cny, "GBP": usd_cny / gbp_usd}
    except:
        return {"USD": 7.2, "GBP": 9.0}

COEFF = {"US": 1.38 * 2, "UK": 1.34 * 2}
SYM   = {"US": "$", "UK": "£"}

def calc_sell(markup_rmb, fx, mode, zi, zs):
    import math
    raw = (markup_rmb + zi + zs) * COEFF[mode] / fx
    return math.ceil(raw) - 0.01  # .99结尾

def extract_number_and_unit(val):
    if val is None: return None, None, None
    s = str(val).strip()
    m = re.match(r'^([^\d]*)(\d+\.?\d*)\s*([a-zA-Z"\'°]*)$',
                 s.replace(' ', '').replace(':', ''))
    if m:
        prefix = re.sub(r'[:\s]', '', m.group(1)).strip()
        return float(m.group(2)), m.group(3).upper() or "", prefix
    return None, None, None

def read_data(filepath, n_dims, fx, mode, zi, zs):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    start_col, start_row, markup_col = 0, 0, n_dims
    for idx, row in enumerate(rows):
        first = next((i for i, v in enumerate(row) if v is not None), None)
        if first is None: continue
        start_col = first
        markup_col = start_col + n_dims
        try:
            float(row[markup_col]); start_row = idx; break
        except (TypeError, ValueError, IndexError):
            continue
    is_text = isinstance(rows[start_row][start_col], str)
    data, units, prefixes = [], [None]*n_dims, [None]*n_dims
    for row in rows[start_row:]:
        try:
            dims = []
            for i in range(n_dims):
                cell = row[start_col + i]
                if is_text:
                    num, unit, prefix = extract_number_and_unit(cell)
                    if num is None: raise ValueError
                    if units[i] is None and unit: units[i] = unit
                    if prefixes[i] is None and prefix: prefixes[i] = prefix
                    dims.append(num)
                else:
                    dims.append(float(cell))
            markup = float(row[markup_col])
            sell   = calc_sell(markup, fx, mode, zi, zs)
            data.append(tuple(dims) + (markup, sell))
        except (TypeError, ValueError, IndexError):
            continue
    if not data:
        raise ValueError("未读取到有效数据，请检查变量列数配置")
    dim_names  = [prefixes[i] or chr(65+i) for i in range(n_dims)]
    dim_values = [sorted(set(d[i] for d in data)) for i in range(n_dims)]
    return data, dim_values, units, dim_names

def optimize(data, dim_values, n_dims, min_ratio):
    offsets = [1]
    for dv in dim_values: offsets.append(offsets[-1] + len(dv) - 1)
    n_vars = offsets[-1]
    y = np.array([d[-1] for d in data])
    def pred_one(x, row):
        t = x[0]
        for i in range(n_dims):
            j = dim_values[i].index(row[i])
            if j > 0: t += x[offsets[i]+j-1]
        return t
    def pred_all(x): return np.array([pred_one(x, r) for r in data])
    cons = [{'type':'ineq','fun': lambda x,r=row,p=float(price): pred_one(x,r)-min_ratio*p}
            for row, price in zip(data, y)]
    x0 = np.ones(n_vars); x0[0] = y.min() * min_ratio
    res = minimize(lambda x: np.sum((pred_all(x)-y)**2), x0,
                   method='SLSQP', bounds=[(0,None)]*n_vars,
                   constraints=cons, options={'ftol':1e-9,'maxiter':10000})
    return res.x, pred_all(res.x), offsets, y

def write_output(path, n_dims, dim_values, dim_names, units,
                 x, offsets, data, pred, y, min_ratio, mode):
    sym = SYM[mode]
    wb  = Workbook()
    hf = PatternFill("solid", start_color="2C3E50")
    sf = PatternFill("solid", start_color="EBF5FB")
    gf = PatternFill("solid", start_color="D5F5E3")
    yf = PatternFill("solid", start_color="FEF9E7")
    rf = PatternFill("solid", start_color="FADBD8")
    cf = PatternFill("solid", start_color="BDC3C7")
    bw   = XFont(bold=True, color="FFFFFF", name="Arial", size=10)
    bold = XFont(bold=True, name="Arial", size=10)
    norm = XFont(name="Arial", size=10)
    redf = XFont(name="Arial", size=10, color="E74C3C")
    ctr  = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    def sc(cell, font=norm, fill=None, border=thin):
        cell.font=font; cell.alignment=ctr; cell.border=border
        if fill: cell.fill=fill
    def fmtv(val, unit): return f"{val}{unit or ''}"
    ws1 = wb.active; ws1.title = "各维度加价表"
    r = 1
    ws1.merge_cells(f"A{r}:C{r}")
    sc(ws1.cell(r,1,f"Amazon 定制加价拆分  [{mode}站]"),
       font=XFont(bold=True,size=13,name="Arial",color="FFFFFF"), fill=hf)
    ws1.row_dimensions[r].height = 26; r += 2
    ws1.merge_cells(f"A{r}:C{r}")
    sc(ws1.cell(r,1,"【基础价格】"), font=bold, fill=sf); r+=1
    for c,h in [(1,"项目"),(2,f"金额({sym})"),(3,"备注")]:
        sc(ws1.cell(r,c,h), font=bold, fill=cf)
    r+=1
    import math as _math
    base_price = _math.ceil(x[0]) - 0.01  # .99结尾
    sc(ws1.cell(r,1,"基础价格"))
    ws1.cell(r,2,base_price); ws1.cell(r,2).number_format=f'"{sym}"#,##0.00'
    sc(ws1.cell(r,2)); sc(ws1.cell(r,3,"所有组合共享，叠加各维度增量")); r+=2
    for i in range(n_dims):
        unit = units[i] or ""
        ws1.merge_cells(f"A{r}:C{r}")
        sc(ws1.cell(r,1,f"【{dim_names[i]} 增量】"), font=bold, fill=sf); r+=1
        for c,h in [(1,f"值({unit})"if unit else"值"),(2,f"增量({sym})"),(3,"备注")]:
            sc(ws1.cell(r,c,h), font=bold, fill=cf)
        r+=1
        for j,val in enumerate(dim_values[i]):
            inc = 0.0 if j==0 else round(x[offsets[i]+j-1])  # 增量取整数
            sc(ws1.cell(r,1,fmtv(val,unit)))
            ws1.cell(r,2,inc); ws1.cell(r,2).number_format=f'"{sym}"#,##0.00'
            sc(ws1.cell(r,2))
            sc(ws1.cell(r,3,"基准档"if j==0 else"增量")); r+=1
        r+=1
    ws1.merge_cells(f"A{r}:C{r}")
    sc(ws1.cell(r,1,f"约束：拆分价 ≥ 原价×{int(min_ratio*100)}%"), font=redf)
    for col,w in [("A",18),("B",14),("C",28)]: ws1.column_dimensions[col].width=w
    ws2 = wb.create_sheet("验证对比表")
    headers = dim_names+["加价(¥)",f"前台售价({sym})",f"拆分价({sym})","差额","误差%","状态"]
    for c,h in enumerate(headers,1): sc(ws2.cell(1,c,h), font=bw, fill=hf)
    errs = (pred-y)/y*100; thr = -(1-min_ratio)*100
    for ri,(row_data,p,e) in enumerate(zip(data,pred,errs),2):
        for c in range(n_dims): sc(ws2.cell(ri,c+1,fmtv(row_data[c],units[c])))
        sc(ws2.cell(ri,n_dims+1,round(row_data[n_dims],2)))
        ws2.cell(ri,n_dims+1).number_format='"¥"#,##0.00'
        sv = row_data[n_dims+1]
        for col,val,fs in [(n_dims+2,round(sv,2),f'"{sym}"#,##0.00'),
                            (n_dims+3,round(p,2),f'"{sym}"#,##0.00'),
                            (n_dims+4,round(p-sv,2),f'"{sym}"#,##0.00')]:
            ws2.cell(ri,col,val).number_format=fs; sc(ws2.cell(ri,col))
        sc(ws2.cell(ri,n_dims+5,round(e,1)))
        if e<thr-0.05:   st,fi="⚠️触红线",rf
        elif e<0:        st,fi=f"↓低{abs(e):.1f}%",yf
        else:            st,fi=f"↑高{e:.1f}%",gf
        sc(ws2.cell(ri,n_dims+6,st),fill=fi)
    for c in range(1,n_dims+7):
        ws2.column_dimensions[ws2.cell(1,c).column_letter].width=14
    ws3 = wb.create_sheet("误差统计")
    for c,h in enumerate(["统计项","数值"],1): sc(ws3.cell(1,c,h),font=bw,fill=hf)
    stats=[("总组合数",len(data)),("低于前台售价",int((errs<0).sum())),
           (f"触碰{int(min_ratio*100)}%红线",int((errs<thr-0.05).sum())),
           ("高于前台售价",int((errs>0).sum())),("最大低于",f"{errs.min():.1f}%"),
           ("最大高于",f"{errs.max():.1f}%"),("平均绝对误差",f"{np.abs(errs).mean():.1f}%")]
    for ri,(k,v) in enumerate(stats,2): sc(ws3.cell(ri,1,k)); sc(ws3.cell(ri,2,v))
    ws3.column_dimensions["A"].width=24; ws3.column_dimensions["B"].width=16
    wb.save(path)
    return errs

# ============================================================
# Worker 线程
# ============================================================

class Worker(QThread):
    done    = pyqtSignal(dict)
    error   = pyqtSignal(str)

    def __init__(self, filepath, n_dims, fx, mode, zi, zs, min_ratio):
        super().__init__()
        self.filepath  = filepath
        self.n_dims    = n_dims
        self.fx        = fx
        self.mode      = mode
        self.zi        = zi
        self.zs        = zs
        self.min_ratio = min_ratio

    def run(self):
        try:
            data, dim_values, units, dim_names = read_data(
                self.filepath, self.n_dims, self.fx, self.mode, self.zi, self.zs)
            x, pred, offsets, y = optimize(data, dim_values, self.n_dims, self.min_ratio)
            errs = (pred - y) / y * 100
            folder   = os.path.join(os.path.dirname(self.filepath), "拆分结果")
            os.makedirs(folder, exist_ok=True)
            basename = os.path.basename(self.filepath).replace(".xlsx", "")
            out_path = os.path.join(folder, f"{basename}_拆分结果_{self.mode}.xlsx")
            write_output(out_path, self.n_dims, dim_values, dim_names, units,
                         x, offsets, data, pred, y, self.min_ratio, self.mode)
            thr = -(1 - self.min_ratio) * 100
            self.done.emit({
                "count":      len(data),
                "avg_err":    f"{np.abs(errs).mean():.1f}%",
                "max_low":    f"{errs.min():.1f}%",
                "max_hi":     f"{errs.max():.1f}%",
                "outfile":    os.path.basename(out_path),
                "outfile_path": out_path,
                "red":        int((errs < thr - 0.05).sum()),
            })
        except Exception as e:
            self.error.emit(str(e))

# ============================================================
# UI 组件
# ============================================================

STYLESHEET = """
QMainWindow, QWidget#central {
    background: #EEEEF0;
}

/* 卡片 */
QFrame#card {
    background: white;
    border-radius: 12px;
    border: 1px solid #D8D8D8;
}

/* 标题栏 */
QFrame#titlebar {
    background: white;
    border-bottom: 1px solid #D8D8D8;
}

/* 标题文字 */
QLabel#title {
    font-family: 'Segoe UI';
    font-size: 16px;
    font-weight: 600;
    color: #1A1A1A;
}

QLabel#section {
    font-family: 'Segoe UI';
    font-size: 10px;
    color: #888888;
    font-weight: 600;
    letter-spacing: 1px;
}

QLabel#result_label {
    font-family: 'Segoe UI';
    font-size: 12px;
    color: #444444;
}

QLabel#result_value {
    font-family: 'Segoe UI';
    font-size: 13px;
    font-weight: 600;
    color: #1A1A1A;
}

QLabel#rate_label {
    font-family: 'Segoe UI';
    font-size: 11px;
    color: #888888;
}

/* 输入框 */
QLineEdit {
    font-family: 'Segoe UI';
    font-size: 13px;
    color: #1A1A1A;
    background: #F7F7F7;
    border: 1.5px solid #CCCCCC;
    border-radius: 6px;
    padding: 5px 10px;
    selection-background-color: #CCE4F7;
}
QLineEdit:focus {
    border: 1.5px solid #0067C0;
    background: white;
}

/* 主按钮 */
QPushButton#run_btn {
    font-family: 'Segoe UI';
    font-size: 14px;
    font-weight: 700;
    color: white;
    background: #0067C0;
    border: none;
    border-radius: 8px;
    padding: 10px 28px;
    min-height: 42px;
}
QPushButton#run_btn:hover   { background: #1478CC; }
QPushButton#run_btn:pressed { background: #005294; }
QPushButton#run_btn:disabled { background: #A0C4E8; color: #E0EEF8; }

/* 次按钮 */
QPushButton#clear_btn {
    font-family: 'Segoe UI';
    font-size: 13px;
    font-weight: 500;
    color: #2A2A2A;
    background: #E8E8E8;
    border: 1.5px solid #C8C8C8;
    border-radius: 8px;
    padding: 10px 22px;
    min-height: 42px;
}
QPushButton#clear_btn:hover  { background: #DCDCDC; }
QPushButton#clear_btn:pressed { background: #CCCCCC; }

/* 模式切换 */
QPushButton#mode_btn {
    font-family: 'Segoe UI';
    font-size: 12px;
    font-weight: 500;
    color: #555555;
    background: transparent;
    border: none;
    border-radius: 6px;
    padding: 5px 14px;
    min-height: 30px;
}
QPushButton#mode_btn:hover { background: #E8E8E8; }
QPushButton#mode_btn_active {
    font-family: 'Segoe UI';
    font-size: 12px;
    font-weight: 700;
    color: #0067C0;
    background: #DCF0FF;
    border: none;
    border-radius: 6px;
    padding: 5px 14px;
    min-height: 30px;
}

/* 进度条 */
QProgressBar {
    background: #D8D8D8;
    border: none;
    border-radius: 3px;
    height: 6px;
    text-align: center;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #0067C0, stop:1 #40A0E0);
    border-radius: 3px;
}
"""

class DropZone(QFrame):
    file_dropped = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setObjectName("dropzone")
        self.setAcceptDrops(True)
        self.setMinimumHeight(130)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._filepath = ""
        self._normal_style = """
            QFrame#dropzone {
                background: #FAFAFA;
                border: 2px dashed #D0D0D0;
                border-radius: 10px;
            }
            QFrame#dropzone:hover {
                border-color: #0067C0;
                background: #F0F7FF;
            }
        """
        self._active_style = """
            QFrame#dropzone {
                background: #E8F2FB;
                border: 2px dashed #0067C0;
                border-radius: 10px;
            }
        """
        self.setStyleSheet(self._normal_style)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(6)

        self.icon = QLabel("📂")
        self.icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.icon.setStyleSheet("font-size: 28px; border: none; background: transparent;")

        self.hint = QLabel("拖入 Excel 文件，或点击选择")
        self.hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.hint.setStyleSheet("""
            font-family: 'Segoe UI'; font-size: 12px;
            color: #8A8A8A; border: none; background: transparent;
        """)

        self.filename = QLabel("")
        self.filename.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.filename.setStyleSheet("""
            font-family: 'Segoe UI'; font-size: 12px; font-weight: 600;
            color: #0067C0; border: none; background: transparent;
        """)

        layout.addWidget(self.icon)
        layout.addWidget(self.hint)
        layout.addWidget(self.filename)

    def mousePressEvent(self, event):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx)")
        if path: self._set_file(path)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(self._active_style)

    def dragLeaveEvent(self, event):
        self.setStyleSheet(self._normal_style)

    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet(self._normal_style)
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            self._set_file(path)

    def _set_file(self, path):
        if not path.lower().endswith(".xlsx"):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "格式错误", "请选择 .xlsx 文件")
            return
        self._filepath = path
        self.icon.setText("✅")
        self.hint.setText("已选择：")
        self.filename.setText(os.path.basename(path))
        self.file_dropped.emit(path)

    def get_path(self): return self._filepath

    def reset(self):
        self._filepath = ""
        self.icon.setText("📂")
        self.hint.setText("拖入 Excel 文件，或点击选择")
        self.filename.setText("")


class ResultRow(QWidget):
    def __init__(self, label, value_color="#1A1A1A"):
        super().__init__()
        self.setStyleSheet("background: transparent;")
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 2, 0, 2)

        lbl = QLabel(label)
        lbl.setObjectName("result_label")
        lbl.setFixedWidth(70)

        self.val = QLabel("—")
        self.val.setObjectName("result_value")
        self.val.setStyleSheet(
            f"font-family: 'Segoe UI'; font-size: 13px; "
            f"font-weight: 500; color: {value_color};")

        layout.addWidget(lbl)
        layout.addWidget(self.val)
        layout.addStretch()

    def set(self, text): self.val.setText(text)


def make_card():
    f = QFrame()
    f.setObjectName("card")
    shadow = QGraphicsDropShadowEffect()
    shadow.setBlurRadius(20)
    shadow.setOffset(0, 2)
    shadow.setColor(QColor(0, 0, 0, 18))
    f.setGraphicsEffect(shadow)
    return f

# ============================================================
# 主窗口
# ============================================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("定制尺寸加价拆分")
        self.setFixedSize(760, 560)
        self.rates  = {"US": None, "UK": None}
        self.mode   = "US"
        self.worker = None
        self._build()
        self._fetch_rates()

    def _build(self):
        central = QWidget()
        central.setObjectName("central")
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ── 标题栏 ──
        titlebar = QFrame()
        titlebar.setObjectName("titlebar")
        titlebar.setFixedHeight(56)
        tb_layout = QHBoxLayout(titlebar)
        tb_layout.setContentsMargins(20, 0, 20, 0)

        icon = QLabel("📦")
        icon.setStyleSheet("font-size: 20px;")
        title = QLabel("定制尺寸加价拆分")
        title.setObjectName("title")

        tb_layout.addWidget(icon)
        tb_layout.addSpacing(8)
        tb_layout.addWidget(title)
        tb_layout.addStretch()

        # 模式切换
        self.mode_btns = {}
        mode_frame = QWidget()
        mode_frame.setStyleSheet("background: transparent;")
        mf_layout = QHBoxLayout(mode_frame)
        mf_layout.setContentsMargins(0,0,0,0)
        mf_layout.setSpacing(4)
        for label, val in [("🇺🇸  美国", "US"), ("🇬🇧  英国", "UK")]:
            btn = QPushButton(label)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _, v=val: self._set_mode(v))
            self.mode_btns[val] = btn
            mf_layout.addWidget(btn)
        tb_layout.addWidget(mode_frame)

        # 汇率
        self.rate_label = QLabel("正在获取汇率…")
        self.rate_label.setObjectName("rate_label")
        tb_layout.addSpacing(16)
        tb_layout.addWidget(self.rate_label)

        main_layout.addWidget(titlebar)

        # 分割线
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: #EBEBEB;")
        main_layout.addWidget(sep)

        # ── 主体 ──
        body = QWidget()
        body.setStyleSheet("background: #F5F5F5;")
        body_layout = QHBoxLayout(body)
        body_layout.setContentsMargins(20, 20, 20, 20)
        body_layout.setSpacing(16)

        # ── 左列 ──
        left = QVBoxLayout()
        left.setSpacing(12)

        # 文件卡片
        file_card = make_card()
        file_layout = QVBoxLayout(file_card)
        file_layout.setContentsMargins(16, 14, 16, 14)
        file_layout.setSpacing(8)
        sec1 = QLabel("EXCEL 文件")
        sec1.setObjectName("section")
        self.drop_zone = DropZone()
        file_layout.addWidget(sec1)
        file_layout.addWidget(self.drop_zone)
        left.addWidget(file_card)

        # 参数卡片
        param_card = make_card()
        param_layout = QVBoxLayout(param_card)
        param_layout.setContentsMargins(16, 14, 16, 14)
        param_layout.setSpacing(10)
        sec2 = QLabel("参数设置")
        sec2.setObjectName("section")
        param_layout.addWidget(sec2)

        self.inputs = {}
        params = [
            [("变量列数", "n_dims", "3", "列"), ("智赢 zi",  "zi",  "10", "")],
            [("智赢 zs",  "zs",    "10", ""), ("最低比例", "min_ratio", "85", "%")],
        ]
        for row_params in params:
            row_w = QWidget(); row_w.setStyleSheet("background: transparent;")
            row_l = QHBoxLayout(row_w)
            row_l.setContentsMargins(0,0,0,0)
            row_l.setSpacing(16)
            for lbl_text, key, default, unit in row_params:
                g = QWidget(); g.setStyleSheet("background: transparent;")
                gl = QHBoxLayout(g); gl.setContentsMargins(0,0,0,0); gl.setSpacing(6)
                lbl = QLabel(lbl_text)
                lbl.setStyleSheet("font-family:'Segoe UI'; font-size:12px; color:#6E6E6E;")
                lbl.setFixedWidth(62)
                entry = QLineEdit(default)
                entry.setFixedWidth(64)
                self.inputs[key] = entry
                gl.addWidget(lbl); gl.addWidget(entry)
                if unit:
                    u = QLabel(unit)
                    u.setStyleSheet("font-family:'Segoe UI'; font-size:11px; color:#AAAAAA;")
                    gl.addWidget(u)
                row_l.addWidget(g)
            row_l.addStretch()
            param_layout.addWidget(row_w)
        left.addWidget(param_card)

        # 按钮行
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self.run_btn = QPushButton("▶   运  行")
        self.run_btn.setObjectName("run_btn")
        self.run_btn.setMinimumHeight(44)
        self.run_btn.setMinimumWidth(160)
        self.run_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.run_btn.clicked.connect(self._run)
        self.run_btn.setStyleSheet("""
            QPushButton {
                font-family: 'Segoe UI'; font-size: 14px; font-weight: 700;
                color: white; background: #0067C0;
                border: none; border-radius: 8px; padding: 10px 28px;
            }
            QPushButton:hover   { background: #1478CC; }
            QPushButton:pressed { background: #005294; }
            QPushButton:disabled { background: #A0C4E8; color: #E0EEF8; }
        """)
        clear_btn = QPushButton("清  空")
        clear_btn.setObjectName("clear_btn")
        clear_btn.setMinimumHeight(44)
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.clicked.connect(self._clear)
        clear_btn.setStyleSheet("""
            QPushButton {
                font-family: 'Segoe UI'; font-size: 13px; font-weight: 500;
                color: #2A2A2A; background: #E8E8E8;
                border: 1.5px solid #C8C8C8; border-radius: 8px; padding: 10px 22px;
            }
            QPushButton:hover  { background: #DCDCDC; }
            QPushButton:pressed { background: #CCCCCC; }
        """)
        btn_row.addWidget(self.run_btn, 3)
        btn_row.addWidget(clear_btn, 1)
        left.addLayout(btn_row)
        left.addStretch()

        # ── 右列 ──
        right = QVBoxLayout()
        result_card = make_card()
        result_card.setMinimumWidth(310)
        rc_layout = QVBoxLayout(result_card)
        rc_layout.setContentsMargins(20, 16, 20, 16)
        rc_layout.setSpacing(4)

        sec3 = QLabel("运行结果")
        sec3.setObjectName("section")
        rc_layout.addWidget(sec3)
        rc_layout.addSpacing(6)

        self.rows = {
            "status":  ResultRow("状 态"),
            "count":   ResultRow("组合数"),
            "avg_err": ResultRow("平均误差"),
            "max_low": ResultRow("最大低于", "#C05A00"),
            "max_hi":  ResultRow("最大高于", "#0F7B0F"),
        }
        for row in self.rows.values():
            rc_layout.addWidget(row)

        # 可点击的输出文件行
        outfile_row = QWidget(); outfile_row.setStyleSheet("background: transparent;")
        outfile_layout = QHBoxLayout(outfile_row)
        outfile_layout.setContentsMargins(0, 2, 0, 2)
        outfile_lbl = QLabel("输出文件")
        outfile_lbl.setObjectName("result_label")
        outfile_lbl.setFixedWidth(70)
        self.outfile_link = QLabel("—")
        self.outfile_link.setStyleSheet(
            "font-family:'Segoe UI'; font-size:13px; font-weight:600; "
            "color:#0067C0; text-decoration: underline; cursor: pointer;")
        self.outfile_link.setCursor(Qt.CursorShape.PointingHandCursor)
        self.outfile_link.mousePressEvent = self._open_outfile
        self._outfile_path = ""
        outfile_layout.addWidget(outfile_lbl)
        outfile_layout.addWidget(self.outfile_link)
        outfile_layout.addStretch()
        rc_layout.addWidget(outfile_row)

        rc_layout.addSpacing(10)

        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setFixedHeight(5)
        self.progress.setTextVisible(False)
        self.progress.setVisible(False)
        rc_layout.addWidget(self.progress)

        self.warn_label = QLabel("")
        self.warn_label.setStyleSheet(
            "font-family:'Segoe UI'; font-size:11px; color:#C42B1C; background:transparent;")
        self.warn_label.setWordWrap(True)
        rc_layout.addWidget(self.warn_label)

        rc_layout.addStretch()
        right.addWidget(result_card)

        body_layout.addLayout(left, 5)
        body_layout.addLayout(right, 4)
        main_layout.addWidget(body)
        self._set_mode("US")

    def _set_mode(self, val):
        self.mode = val
        for v, btn in self.mode_btns.items():
            if v == val:
                btn.setObjectName("mode_btn_active")
            else:
                btn.setObjectName("mode_btn")
            btn.setStyle(btn.style())
        rate = self.rates.get(val)
        sym  = "USD" if val=="US" else "GBP"
        self.rate_label.setText(
            f"1 {sym} = ¥{rate:.4f}" if rate else f"获取 {sym} 汇率中…")

    def _fetch_rates(self):
        from PyQt6.QtCore import QTimer
        def _f():
            r = get_all_rates()
            self.rates["US"] = r["USD"]
            self.rates["UK"] = r["GBP"]
            QTimer.singleShot(0, lambda: self._set_mode(self.mode))
        threading.Thread(target=_f, daemon=True).start()

    def _clear(self):
        self.drop_zone.reset()
        defaults = {"n_dims":"3","zi":"10","zs":"10","min_ratio":"85"}
        for k,v in defaults.items(): self.inputs[k].setText(v)
        for row in self.rows.values(): row.set("—")
        self.outfile_link.setText("—")
        self._outfile_path = ""
        self.warn_label.setText("")
        self.progress.setVisible(False)

    def _run(self):
        filepath = self.drop_zone.get_path()
        if not filepath:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "错误", "请先选择或拖入 Excel 文件")
            return
        try:
            n_dims    = int(self.inputs["n_dims"].text())
            zi        = float(self.inputs["zi"].text())
            zs        = float(self.inputs["zs"].text())
            min_ratio = float(self.inputs["min_ratio"].text()) / 100
        except ValueError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "输入错误", "参数请填写数字")
            return

        fx = self.rates.get(self.mode) or (7.2 if self.mode=="US" else 9.0)

        self.run_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.rows["status"].set("⏳ 计算中…")
        self.warn_label.setText("")

        self.worker = Worker(filepath, n_dims, fx, self.mode, zi, zs, min_ratio)
        self.worker.done.connect(self._on_done)
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _open_outfile(self, event=None):
        if self._outfile_path and os.path.exists(self._outfile_path):
            os.startfile(self._outfile_path)

    def _on_done(self, result):
        self.run_btn.setEnabled(True)
        self.progress.setVisible(False)
        self.rows["status"].set("✅ 完成")
        self.rows["count"].set(f"{result['count']} 个组合")
        self.rows["avg_err"].set(result["avg_err"])
        self.rows["max_low"].set(result["max_low"])
        self.rows["max_hi"].set(result["max_hi"])
        self._outfile_path = result["outfile_path"]
        self.outfile_link.setText(result["outfile"])
        if result["red"]:
            self.warn_label.setText(f"⚠  {result['red']} 个组合触碰红线，请检查验证对比表")

    def _on_error(self, msg):
        self.run_btn.setEnabled(True)
        self.progress.setVisible(False)
        self.rows["status"].set("❌ 出错")
        self.warn_label.setText(f"错误：{msg}")


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)
    # 启用高DPI
    app.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())